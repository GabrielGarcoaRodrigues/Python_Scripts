from pathlib import Path    
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'ratings.xlsx'

workbook = Workbook()

sheetName = 'Ratings'
workbook.create_sheet(sheetName) #Cria planilha nova

workbook.remove(workbook.active) #Remove a sheet atual
worksheet: Worksheet = workbook.active #Seleciona a sheet criada

#criando cabeçalhos
worksheet.cell(row=1, column=1, value='Name')
worksheet.cell(row=1, column=2, value='Age')
worksheet.cell(row=1, column=3, value='Rating')

students = [
    #nome, idade, nota
    ['João',15,5.7],
    ['Maria',16,6.2],
    ['José',15,5.7],
    ['Ana',16,6.2],
]

# for i in range(2, 6):
#     for j in range(1, 4):
#         # print('Linha', i , 'Coluna', j)
#         worksheet.cell(row=i, column=j, value=students[i-2][j-1])

for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)